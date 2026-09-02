import io
import os
import re
import csv
import uuid
import zipfile
from datetime import datetime
from django.conf import settings
from django.core.mail import EmailMessage
from django.utils import timezone
from .models import Certificate, generate_next_certificate_number
from .pdf_generator import generate_certificate_pdf


def send_certificate_email(certificate, pdf_bytes: bytes, base_url="http://localhost:5173") -> tuple[bool, str | None]:
    """
    Send an official digital certificate email to recipient with vector PDF attached.
    """
    subject = f"Official Certificate Awarded: {certificate.title} ({certificate.certificate_number})"
    verify_url = f"{base_url.rstrip('/')}/verify/{certificate.certificate_number}"
    
    body = f"""Dear {certificate.recipient_name},

Congratulations! You have been awarded an official digital credential:

Certificate Title: {certificate.title}
Achievement: {certificate.achievement or 'Distinction & Outstanding Performance'}
Serial ID: {certificate.certificate_number}
Issuing Organization: {certificate.organization_name}
Date of Issue: {certificate.created_at.strftime('%B %d, %Y') if certificate.created_at else 'Official Record'}

Your high-resolution vector certificate is attached to this email as a PDF.

Verification & Authenticity:
You, academic institutions, and employers can verify the authenticity, issue date, and validity of this certificate at any time via the official CertiGen Ledger:
{verify_url}

Authorized Signatory:
{certificate.signatory_name}
{certificate.signatory_title}
{certificate.organization_name}

---
CertiGen Digital Certificate Verification Authority
"""

    try:
        from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', 'CertiGen <no-reply@certigen.io>')
        email = EmailMessage(
            subject=subject,
            body=body,
            from_email=from_email,
            to=[certificate.recipient_email],
        )
        safe_name = re.sub(r'[^a-zA-Z0-9_\-]', '_', certificate.recipient_name)
        filename = f"{certificate.certificate_number}_{safe_name}.pdf"
        email.attach(filename, pdf_bytes, 'application/pdf')
        email.send(fail_silently=False)
        return True, None
    except Exception as e:
        return False, str(e)


def parse_recipients_file(file_obj, filename: str) -> list[dict]:
    """
    Parse uploaded Excel (.xlsx, .xls) or CSV sheet and return normalized list of recipient dicts.
    """
    recipients = []
    lower_filename = filename.lower()
    
    if lower_filename.endswith('.csv'):
        # Parse CSV
        content = file_obj.read()
        if isinstance(content, bytes):
            content = content.decode('utf-8-sig', errors='replace')
        
        reader = csv.DictReader(io.StringIO(content))
        for row in reader:
            normalized = _normalize_row_keys(row)
            if normalized.get('recipient_name') or normalized.get('name'):
                recipients.append(_standardize_recipient_fields(normalized))
                
    elif lower_filename.endswith(('.xlsx', '.xls')):
        # Parse Excel using openpyxl
        import openpyxl
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        sheet = wb.active
        
        headers = []
        for col_idx, cell in enumerate(sheet.iter_rows(min_row=1, max_row=1, values_only=True)):
            headers = [str(c or '').strip().lower().replace(' ', '_') for c in cell]
            break
            
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            row_dict = {}
            for col_idx, value in enumerate(row):
                if col_idx < len(headers):
                    key = headers[col_idx]
                    row_dict[key] = str(value or '').strip()
            
            normalized = _normalize_row_keys(row_dict)
            if normalized.get('recipient_name') or normalized.get('name'):
                recipients.append(_standardize_recipient_fields(normalized))
                
    return recipients


def _normalize_row_keys(row: dict) -> dict:
    normalized = {}
    for k, v in row.items():
        clean_key = str(k or '').strip().lower().replace(' ', '_').replace('-', '_')
        clean_val = str(v or '').strip()
        normalized[clean_key] = clean_val
    return normalized


def _standardize_recipient_fields(row: dict) -> dict:
    name = (
        row.get('recipient_name') or 
        row.get('student_name') or 
        row.get('full_name') or 
        row.get('name') or 
        ''
    )
    email = (
        row.get('recipient_email') or 
        row.get('student_email') or 
        row.get('email_address') or 
        row.get('email') or 
        ''
    )
    achievement = (
        row.get('achievement') or 
        row.get('honor') or 
        row.get('award') or 
        ''
    )
    rank = row.get('rank') or row.get('position') or ''
    duration = row.get('duration') or row.get('period') or ''
    instructor = row.get('instructor') or row.get('mentor') or row.get('teacher') or ''
    team_name = row.get('team_name') or row.get('team') or ''
    role = row.get('role') or row.get('designation') or ''
    hours = row.get('hours') or ''

    return {
        'recipient_name': name,
        'recipient_email': email,
        'achievement': achievement,
        'rank': rank,
        'duration': duration,
        'instructor': instructor,
        'team_name': team_name,
        'role': role,
        'hours': hours,
    }


def generate_bulk_certificates_and_zip(
    user,
    template,
    recipients: list[dict],
    common_data: dict,
    send_emails: bool = False,
    base_url: str = "http://localhost:5173"
) -> dict:
    """
    Generate certificates for every recipient in list,
    optionally email each one with PDF attached,
    and bundle all PDFs into a single downloadable ZIP archive.
    """
    batch_id = str(uuid.uuid4())
    batches_dir = os.path.join(settings.MEDIA_ROOT, 'batches')
    os.makedirs(batches_dir, exist_ok=True)
    
    timestamp_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    zip_filename = f"CertiGen_Batch_{timestamp_str}_{batch_id[:8]}.zip"
    zip_filepath = os.path.join(batches_dir, zip_filename)

    issued_certificates = []
    emails_sent_count = 0
    errors = []

    # Prepare in-memory or on-disk ZIP archive
    with zipfile.ZipFile(zip_filepath, 'w', compression=zipfile.ZIP_DEFLATED) as zipf:
        for idx, rec in enumerate(recipients):
            name = rec.get('recipient_name', '').strip()
            email = rec.get('recipient_email', '').strip()
            if not name:
                continue

            # Achievement fallback
            rec_achievement = rec.get('achievement') or common_data.get('achievement') or 'Distinction'
            title = common_data.get('title') or common_data.get('event_name') or 'Certificate of Achievement'
            org_name = common_data.get('organization_name') or 'CertiGen Platform'
            signatory_name = common_data.get('signatory_name') or 'Dr. Rajesh Kumar'
            signatory_title = common_data.get('signatory_title') or 'Dean of Academic Affairs'

            # Dynamic wording resolution
            wording_pattern = getattr(template, 'wording_pattern', '') or 'for outstanding achievement in {{EVENT_NAME}}'
            resolved_wording = (
                wording_pattern
                .replace('{{STUDENT_NAME}}', name)
                .replace('{{NAME}}', name)
                .replace('{{EVENT_NAME}}', title)
                .replace('{{COURSE_NAME}}', title)
                .replace('{{ACHIEVEMENT}}', rec_achievement)
                .replace('{{ORGANIZATION_NAME}}', org_name)
                .replace('{{DATE}}', common_data.get('issue_date') or datetime.now().strftime('%Y-%m-%d'))
                .replace('{{RANK}}', rec.get('rank') or 'Distinction')
                .replace('{{DURATION}}', rec.get('duration') or '8 Weeks')
                .replace('{{INSTRUCTOR}}', rec.get('instructor') or 'Course Mentor')
                .replace('{{TEAM_NAME}}', rec.get('team_name') or 'Team')
                .replace('{{ROLE}}', rec.get('role') or 'Intern')
                .replace('{{HOURS}}', rec.get('hours') or '50')
            )

            # Generate Next ID
            cert_number = generate_next_certificate_number()

            # Metadata
            cert_meta = {
                'rank': rec.get('rank', ''),
                'duration': rec.get('duration', ''),
                'instructor': rec.get('instructor', ''),
                'team_name': rec.get('team_name', ''),
                'role': rec.get('role', ''),
                'hours': rec.get('hours', ''),
                'issue_date': common_data.get('issue_date', datetime.now().strftime('%Y-%m-%d')),
                'primary_color': common_data.get('primary_color') or getattr(template, 'primary_color', '#0f2744'),
                'secondary_color': common_data.get('secondary_color') or getattr(template, 'secondary_color', '#c59b27'),
                'accent_color': common_data.get('accent_color') or getattr(template, 'accent_color', '#e2d19f'),
                'institute_logo_base64': common_data.get('institute_logo_base64', ''),
                'second_signatory_name': common_data.get('second_signatory_name', 'Prof. Vikram Singh'),
                'second_signatory_title': common_data.get('second_signatory_title', 'Director of Certification'),
                'institute_subtitle': common_data.get('institute_subtitle', ''),
                'batch_id': batch_id,
            }

            # Create Certificate model
            cert = Certificate.objects.create(
                certificate_number=cert_number,
                title=title,
                description=resolved_wording,
                template=template,
                recipient_name=name,
                recipient_email=email,
                achievement=rec_achievement,
                organization_name=org_name,
                signatory_name=signatory_name,
                signatory_title=signatory_title,
                metadata=cert_meta,
                issued_by=user,
                status='VALID'
            )

            # Generate PDF bytes
            pdf_bytes = generate_certificate_pdf(cert, base_url=base_url)

            # Add PDF to ZIP
            safe_name = re.sub(r'[^a-zA-Z0-9_\-]', '_', name)
            pdf_archive_filename = f"{cert_number}_{safe_name}.pdf"
            zipf.writestr(pdf_archive_filename, pdf_bytes)

            # Send Email if requested
            email_sent = False
            if send_emails and email:
                email_sent, err_msg = send_certificate_email(cert, pdf_bytes, base_url=base_url)
                if email_sent:
                    emails_sent_count += 1
                else:
                    errors.append(f"Email error for {name} ({email}): {err_msg}")

            issued_certificates.append({
                'id': str(cert.id),
                'certificate_number': cert.certificate_number,
                'recipient_name': cert.recipient_name,
                'recipient_email': cert.recipient_email,
                'achievement': cert.achievement,
                'email_sent': email_sent,
                'pdf_filename': pdf_archive_filename
            })

    return {
        'batch_id': batch_id,
        'total_issued': len(issued_certificates),
        'emails_sent': emails_sent_count,
        'zip_filename': zip_filename,
        'zip_relative_url': f"/media/batches/{zip_filename}",
        'certificates': issued_certificates,
        'errors': errors
    }
